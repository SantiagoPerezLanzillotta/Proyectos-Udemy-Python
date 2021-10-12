#Ejercicio 1: Multiplication Table Make

'''
Create a program multiplicationTable.py that takes a number N from the
command line and creates an N×N multiplication table in an Excel spreadsheet.

For example, when the program is run like this:

py multiplicationTable.py 6
. . . it should create a spreadsheet that looks like:

- A B  C  D  E  F  G   
1   1  2  3  4  5  6
2 1 1  2  3  4  5  6
3 2 2  4  6  8 10 12
4 3 3  6  9 12 15 18
5 4 4  8 12 16 20 24
6 5 5 10 15 20 25 30
7 6 6 12 18 24 30 36

Row 1 and column A should be used for labels and should be in bold.

'''
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

entrada = 6

wb = openpyxl.Workbook() #creo un excel en blanco
#print(wb.worksheets)

sheet = wb['Sheet']

for i in range (1,entrada+1): #Seteo los valores inciales
    sheet.cell(row=i+1,column=1).value = i
    sheet.cell(row=1,column=i+1).value = i
        

#print(sheet['A2'].value)
#print(sheet['A7'].value)

#print(sheet['B1'].value)
#print(sheet['G1'].value)

for f in range(2,entrada+2):     #fila
    for c in range(2,entrada+2): #columna
        stringColumna = openpyxl.utils.get_column_letter(c)
        sheet.cell(row=f,column=c).value = '=' + stringColumna+'1*A' +str(f)
        



wb.save('Multiplication Table Maker.xlsx')




    
    
