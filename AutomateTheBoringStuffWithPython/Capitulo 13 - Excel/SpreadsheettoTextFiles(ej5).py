#Ejercicio 5: Spreadsheet to Text Files

'''
Write a program that performs the tasks of the previous program in reverse
order: the program should open a spreadsheet and write the cells of column A
into one text file, the cells of column B into another text file, and so on.

'''
import openpyxl

excel = input('Ingrese el archivo excel: ')
indiceHoja = int(input('Ingrese de que indice de hoja sacar el texto: '))
wb = openpyxl.load_workbook(excel)

contador = 0
if indiceHoja < len(wb.worksheets):

    sheet = wb.worksheets[indiceHoja]

    for y in range(0,sheet.max_column):
        file = open('File' + str(contador) + '.txt','w')
        for i in range(0,sheet.max_row):
            valor = sheet.cell(row=i+1,column=y+1).value
            if valor != None:
                file.write(str(valor))
                file.write('\n')
        file.close()
        contador+=1
        
    
    
