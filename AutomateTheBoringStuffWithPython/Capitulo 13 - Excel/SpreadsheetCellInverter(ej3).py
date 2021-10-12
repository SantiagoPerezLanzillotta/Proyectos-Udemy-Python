#Ejercicio 3: Spreadsheet Cell Inverter

'''
Write a program to invert the row and column of the cells in the spreadsheet.

For example, the value at row 5, column 3 will be at row 3, column 5 (and vice
versa).


This should be done for all cells in the spreadsheet.
For example, the “before” and “after” spreadsheets would look something

Before:
      A       B
1-  hola     auto
2-  voy     amigos
3-   ir     teclado
4- bailar   gaming
5- loco     comida

After:
    A      B       C         D      E
1- hola   voy      ir      bailar   loco
2- auto   amigos  teclado  gaming  comida


You can write this program by using nested for loops to read the spreadsheet’s
data into a list of lists data structure.

This data structure could have sheetData[x][y] for the cell at column x and
row y.

Then, when writing out the new spreadsheet, use sheetData[y][x] for
the cell at column x and row y.
'''

#sheet.max_row ~~~> The maximum row index containing data (1-based)

import openpyxl 

DATA = [] #esto será lista de listas, donde cada lista es una columna completa
#[ [hola,voy,a,bailar,loco],
#  [con,amigos,intenso,gaming,comida]]

wb = openpyxl.load_workbook('spreedsheetinverter.xlsx') #Nombre del excel a modificar(hay que tenerlo) hay que crear un excel que ya tenga esta tabla a invertir
indiceHoja = int(input('Ingrese que hoja quiere dar vuelta: ' ))
if indiceHoja <= len(wb.worksheets):
    
    sheet = wb.worksheets[indiceHoja]
    #print(sheet.max_row)

    for y in range(1,sheet.max_column+1):  #la ultima está escrita, entonces tengo que sumarle uno para incluirlo
        columnas = []
        for i in range(1,sheet.max_row+1):
            columnas.append(sheet.cell(row=i,column=y).value)
        DATA.append(columnas)

    if len(DATA)>0:     
        largo = len(DATA[0])
        for i in range(0,largo): #Borro todas las columnas y filas
            for y in range(0,len(DATA)):
                sheet.cell(row=i+1,column=y+1).value = None


    if len(DATA)>0:     
        largo = len(DATA[0])
        for i in range(0,len(DATA)): #escribo en los lugares que corresponden a partir de DATA
            for y in range(0,largo):
                sheet.cell(row=i+1,column=y+1).value = DATA[i][y]

               

    wb.save('Inverter.xlsx')
                
            
